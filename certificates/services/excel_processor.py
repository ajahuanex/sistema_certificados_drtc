"""Servicio para procesar archivos Excel de participantes"""
import openpyxl
from typing import Dict, List, Tuple
from datetime import datetime
import re
import logging

logger = logging.getLogger('certificates')


class ExcelProcessorService:
    """Procesa archivos Excel y crea participantes"""
    
    REQUIRED_COLUMNS = [
        'DNI',
        'Nombres y Apellidos',
        'Fecha del Evento',
        'Tipo de Asistente',
        'Nombre del Evento'
    ]
    
    def validate_file(self, file) -> Tuple[bool, List[str]]:
        """
        Valida que el archivo Excel tenga las columnas requeridas
        
        Args:
            file: Archivo Excel a validar
            
        Returns:
            Tuple con (es_válido, lista_de_errores)
        """
        errors = []
        
        try:
            workbook = openpyxl.load_workbook(file, read_only=True)
            sheet = workbook.active
            
            # Obtener la primera fila (encabezados)
            headers = []
            for cell in sheet[1]:
                if cell.value:
                    headers.append(str(cell.value).strip())
            
            # Verificar que todas las columnas requeridas estén presentes
            missing_columns = []
            for required_col in self.REQUIRED_COLUMNS:
                if required_col not in headers:
                    missing_columns.append(required_col)
            
            if missing_columns:
                errors.append(f"Columnas faltantes: {', '.join(missing_columns)}")
            
            workbook.close()
            
        except Exception as e:
            errors.append(f"Error al leer el archivo: {str(e)}")
        
        return (len(errors) == 0, errors)
    
    def _parse_row(self, row, headers: List[str]) -> Dict:
        """
        Extrae datos de una fila Excel
        
        Args:
            row: Fila del Excel
            headers: Lista de encabezados
            
        Returns:
            Diccionario con los datos de la fila
        """
        row_data = {}
        
        for idx, cell in enumerate(row):
            if idx < len(headers):
                header = headers[idx]
                value = cell.value
                
                # Limpiar espacios en blanco
                if isinstance(value, str):
                    value = value.strip()
                
                row_data[header] = value
        
        return row_data
    
    def _validate_row(self, row_data: Dict) -> Tuple[bool, str]:
        """
        Valida los datos de una fila
        
        Args:
            row_data: Diccionario con datos de la fila
            
        Returns:
            Tuple con (es_válido, mensaje_de_error)
        """
        # Validar que todos los campos requeridos estén presentes
        for col in self.REQUIRED_COLUMNS:
            if col not in row_data or not row_data[col]:
                return (False, f"Campo '{col}' está vacío")
        
        # Validar formato de DNI (debe tener entre 1 y 8 dígitos numéricos)
        dni_raw = str(row_data['DNI']).strip()
        # Remover cualquier caracter no numérico
        dni_clean = ''.join(filter(str.isdigit, dni_raw))
        
        if not dni_clean:
            return (False, f"DNI '{dni_raw}' debe contener dígitos numéricos")
        
        if len(dni_clean) > 8:
            return (False, f"DNI '{dni_raw}' no puede tener más de 8 dígitos")
        
        # Validar tipo de asistente
        valid_types = ['ASISTENTE', 'PONENTE', 'ORGANIZADOR']
        attendee_type = str(row_data['Tipo de Asistente']).strip().upper()
        if attendee_type not in valid_types:
            return (False, f"Tipo de asistente '{attendee_type}' no es válido. Debe ser: {', '.join(valid_types)}")
        
        # Validar fecha
        fecha = row_data['Fecha del Evento']
        if isinstance(fecha, str):
            # Intentar parsear diferentes formatos de fecha
            date_formats = ['%d/%m/%Y', '%Y-%m-%d', '%d-%m-%Y']
            fecha_valida = False
            for fmt in date_formats:
                try:
                    datetime.strptime(fecha, fmt)
                    fecha_valida = True
                    break
                except ValueError:
                    continue
            
            if not fecha_valida:
                return (False, f"Formato de fecha inválido: '{fecha}'. Use DD/MM/YYYY")
        elif not isinstance(fecha, datetime):
            return (False, f"Formato de fecha inválido: '{fecha}'")
        
        return (True, "")

    
    def _create_or_update_participant(self, row_data: Dict):
        """
        Crea o actualiza un participante y su evento
        
        Args:
            row_data: Diccionario con datos de la fila
            
        Returns:
            Instancia de Participant creada o actualizada
        """
        from certificates.models import Event, Participant
        
        # Parsear fecha
        fecha = row_data['Fecha del Evento']
        if isinstance(fecha, str):
            # Intentar parsear diferentes formatos
            date_formats = ['%d/%m/%Y', '%Y-%m-%d', '%d-%m-%Y']
            for fmt in date_formats:
                try:
                    fecha = datetime.strptime(fecha, fmt).date()
                    break
                except ValueError:
                    continue
        elif isinstance(fecha, datetime):
            fecha = fecha.date()
        
        # Crear o obtener el evento
        event, created = Event.objects.get_or_create(
            name=row_data['Nombre del Evento'].strip(),
            event_date=fecha,
            defaults={
                'description': ''
            }
        )
        
        if created:
            logger.info(f"Evento creado: {event.name} - {event.event_date}")
        
        # Procesar DNI: asegurar que tenga 8 dígitos con ceros a la izquierda
        dni_raw = str(row_data['DNI']).strip()
        # Remover cualquier caracter no numérico
        dni_clean = ''.join(filter(str.isdigit, dni_raw))
        # Rellenar con ceros a la izquierda hasta 8 dígitos
        dni = dni_clean.zfill(8)
        
        attendee_type = str(row_data['Tipo de Asistente']).strip().upper()
        
        participant, created = Participant.objects.update_or_create(
            dni=dni,
            event=event,
            defaults={
                'full_name': row_data['Nombres y Apellidos'].strip(),
                'attendee_type': attendee_type
            }
        )
        
        if created:
            logger.info(f"Participante creado: {participant.full_name} ({participant.dni})")
        else:
            logger.info(f"Participante actualizado: {participant.full_name} ({participant.dni})")
        
        return participant

    
    def process_excel(self, file, user=None) -> Dict:
        """
        Procesa un archivo Excel completo y crea participantes
        
        Args:
            file: Archivo Excel a procesar
            user: Usuario que realiza la importación (opcional)
            
        Returns:
            Diccionario con resultados: {
                'success_count': int,
                'error_count': int,
                'errors': list
            }
        """
        from certificates.models import AuditLog
        
        # Validar archivo
        is_valid, errors = self.validate_file(file)
        if not is_valid:
            logger.error(f"Archivo inválido: {errors}")
            return {
                'success_count': 0,
                'error_count': 1,
                'errors': errors
            }
        
        # Resetear el puntero del archivo
        file.seek(0)
        
        success_count = 0
        error_count = 0
        errors = []
        
        try:
            workbook = openpyxl.load_workbook(file)
            sheet = workbook.active
            
            # Obtener encabezados
            headers = []
            for cell in sheet[1]:
                if cell.value:
                    headers.append(str(cell.value).strip())
            
            # Procesar cada fila (empezando desde la fila 2)
            for row_idx, row in enumerate(sheet.iter_rows(min_row=2), start=2):
                # Verificar si la fila está vacía
                if all(cell.value is None or str(cell.value).strip() == '' for cell in row):
                    continue
                
                try:
                    # Parsear fila
                    row_data = self._parse_row(row, headers)
                    
                    # Validar fila
                    is_valid, error_msg = self._validate_row(row_data)
                    if not is_valid:
                        error_count += 1
                        errors.append(f"Fila {row_idx}: {error_msg}")
                        logger.warning(f"Fila {row_idx} inválida: {error_msg}")
                        continue
                    
                    # Crear o actualizar participante
                    self._create_or_update_participant(row_data)
                    success_count += 1
                    
                except Exception as e:
                    error_count += 1
                    error_msg = f"Fila {row_idx}: Error al procesar - {str(e)}"
                    errors.append(error_msg)
                    logger.error(error_msg)
            
            workbook.close()
            
            # Registrar en auditoría
            AuditLog.objects.create(
                action_type='IMPORT',
                user=user,
                description=f"Importación de Excel: {success_count} éxitos, {error_count} errores",
                metadata={
                    'success_count': success_count,
                    'error_count': error_count,
                    'total_rows': success_count + error_count
                }
            )
            
            logger.info(f"Importación completada: {success_count} éxitos, {error_count} errores")
            
        except Exception as e:
            error_msg = f"Error al procesar archivo: {str(e)}"
            logger.error(error_msg)
            errors.append(error_msg)
            error_count += 1
        
        return {
            'success_count': success_count,
            'error_count': error_count,
            'errors': errors
        }
