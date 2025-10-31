"""Vistas de administración para certificados"""
from django.views.generic import FormView
from django.contrib.admin.views.decorators import staff_member_required
from django.utils.decorators import method_decorator
from django.contrib import messages
from django.urls import reverse_lazy
from certificates.forms import ExcelImportForm
from certificates.services.excel_processor import ExcelProcessorService


@method_decorator(staff_member_required, name="dispatch")
class ExcelImportView(FormView):
    """Vista para importar participantes desde Excel"""

    template_name = "admin/certificates/excel_import.html"
    form_class = ExcelImportForm
    success_url = reverse_lazy("certificates:import_excel")

    def form_valid(self, form):
        """Procesa el archivo Excel cuando el formulario es válido"""
        excel_file = form.cleaned_data["excel_file"]

        # Procesar archivo con el servicio
        service = ExcelProcessorService()
        result = service.process_excel(excel_file, user=self.request.user)

        # Mostrar mensajes según el resultado
        if result["success_count"] > 0:
            messages.success(
                self.request,
                f"✓ Importación exitosa: {result['success_count']} participantes procesados.",
            )

        if result["error_count"] > 0:
            messages.warning(
                self.request,
                f"⚠ Se encontraron {result['error_count']} errores durante la importación.",
            )

            # Mostrar los primeros 5 errores
            for error in result["errors"][:5]:
                messages.error(self.request, error)

            if len(result["errors"]) > 5:
                messages.info(
                    self.request,
                    f"... y {len(result['errors']) - 5} errores más.",
                )

        # Guardar resultado en la sesión para mostrarlo en el template
        self.request.session["import_result"] = result

        return super().form_valid(form)

    def get_context_data(self, **kwargs):
        """Agrega datos adicionales al contexto"""
        context = super().get_context_data(**kwargs)

        # Recuperar resultado de importación si existe
        import_result = self.request.session.pop("import_result", None)
        if import_result:
            context["import_result"] = import_result

        context["title"] = "Importar Participantes desde Excel"

        return context



@method_decorator(staff_member_required, name="dispatch")
class ExternalCertificateImportView(FormView):
    """Vista para importar certificados externos desde Excel"""

    template_name = "admin/certificates/external_import.html"
    form_class = ExcelImportForm  # Reutilizamos el mismo formulario
    success_url = reverse_lazy("certificates:import_external")

    def get_context_data(self, **kwargs):
        """Agrega contexto adicional al template"""
        context = super().get_context_data(**kwargs)
        context['title'] = 'Importar Certificados Externos'
        return context

    def form_valid(self, form):
        """Procesa el archivo Excel cuando el formulario es válido"""
        import tempfile
        import os
        
        excel_file = form.cleaned_data["excel_file"]

        # Guardar archivo temporalmente si está en memoria
        if hasattr(excel_file, 'temporary_file_path'):
            file_path = excel_file.temporary_file_path()
        else:
            # Archivo en memoria, guardarlo temporalmente
            with tempfile.NamedTemporaryFile(delete=False, suffix='.xlsx') as tmp_file:
                for chunk in excel_file.chunks():
                    tmp_file.write(chunk)
                file_path = tmp_file.name

        try:
            # Procesar archivo con el servicio de importación externa
            from certificates.services.external_certificate_importer import ExternalCertificateImporter
            
            service = ExternalCertificateImporter()
            result = service.import_from_file(file_path)
        finally:
            # Limpiar archivo temporal si fue creado
            if not hasattr(excel_file, 'temporary_file_path'):
                try:
                    os.unlink(file_path)
                except:
                    pass

        # Mostrar mensajes según el resultado
        if result.get('success', False):
            if result["success_count"] > 0:
                messages.success(
                    self.request,
                    f"✓ {result['success_count']} certificados externos importados exitosamente.",
                )
            
            if result["updated_count"] > 0:
                messages.info(
                    self.request,
                    f"ℹ {result['updated_count']} certificados existentes actualizados.",
                )

            if result["error_count"] > 0:
                messages.warning(
                    self.request,
                    f"⚠ Se encontraron {result['error_count']} errores durante la importación.",
                )

                # Mostrar los primeros 5 errores
                for error in result["errors"][:5]:
                    messages.error(self.request, error)

                if len(result["errors"]) > 5:
                    messages.info(
                        self.request,
                        f"... y {len(result['errors']) - 5} errores más.",
                    )
        else:
            messages.error(
                self.request,
                f"✗ Error en la importación: {result.get('error', 'Error desconocido')}"
            )

        # Guardar resultado en la sesión para mostrarlo en el template
        self.request.session["import_result"] = result

        # Registrar en auditoría
        from certificates.models import AuditLog
        AuditLog.objects.create(
            action_type="IMPORT",
            user=self.request.user,
            description=f"Importación de certificados externos: {result['success_count']} éxitos, {result['error_count']} errores",
            metadata=result,
            ip_address=self.request.META.get("REMOTE_ADDR"),
        )

        return super().form_valid(form)



from django.http import HttpResponse
from django.views import View
from django.contrib.admin.views.decorators import staff_member_required
from django.utils.decorators import method_decorator
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from datetime import date


@method_decorator(staff_member_required, name="dispatch")
class DownloadParticipantsTemplateView(View):
    """Vista para descargar plantilla Excel de participantes"""
    
    def get(self, request):
        # Crear workbook
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Participantes"
        
        # Estilos
        header_fill = PatternFill(start_color="1565C0", end_color="1565C0", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF", size=12)
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # Headers
        headers = [
            "DNI",
            "Nombres y Apellidos",
            "Fecha del Evento",
            "Tipo de Asistente",
            "Nombre del Evento"
        ]
        
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num)
            cell.value = header
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = border
        
        # Datos de ejemplo
        examples = [
            ["12345678", "Juan Pérez García", "15/10/2024", "ASISTENTE", "Capacitación en Seguridad Vial"],
            ["87654321", "María López Quispe", "15/10/2024", "PONENTE", "Capacitación en Seguridad Vial"],
            ["11223344", "Carlos Mamani Flores", "15/10/2024", "ORGANIZADOR", "Capacitación en Seguridad Vial"],
        ]
        
        for row_num, example in enumerate(examples, 2):
            for col_num, value in enumerate(example, 1):
                cell = ws.cell(row=row_num, column=col_num)
                cell.value = value
                cell.border = border
                cell.alignment = Alignment(horizontal='left', vertical='center')
        
        # Ajustar anchos de columna
        ws.column_dimensions['A'].width = 12
        ws.column_dimensions['B'].width = 30
        ws.column_dimensions['C'].width = 18
        ws.column_dimensions['D'].width = 18
        ws.column_dimensions['E'].width = 40
        
        # Crear respuesta HTTP
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = 'attachment; filename=plantilla_participantes.xlsx'
        
        wb.save(response)
        return response


@method_decorator(staff_member_required, name="dispatch")
class DownloadExternalCertificatesTemplateView(View):
    """Vista para descargar plantilla Excel de certificados externos"""
    
    def get(self, request):
        # Crear workbook
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Certificados Externos"
        
        # Estilos
        header_fill = PatternFill(start_color="6F42C1", end_color="6F42C1", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF", size=12)
        required_fill = PatternFill(start_color="FFF3CD", end_color="FFF3CD", fill_type="solid")
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # Headers
        headers = [
            "DNI",
            "Nombres y Apellidos",
            "Fecha del Evento",
            "Tipo de Asistente",
            "Nombre del Evento",
            "URL del Certificado",
            "Sistema Externo"
        ]
        
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num)
            cell.value = header
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = border
            
            # Resaltar columna URL del Certificado
            if header == "URL del Certificado":
                cell.fill = required_fill
                cell.font = Font(bold=True, color="E65100", size=12)
        
        # Datos de ejemplo
        examples = [
            [
                "12345678",
                "Juan Pérez García",
                "15/10/2024",
                "ASISTENTE",
                "Capacitación en Seguridad Vial",
                "https://sistema-antiguo.com/certificados/12345678.pdf",
                "Sistema Antiguo v1.0"
            ],
            [
                "87654321",
                "María López Quispe",
                "15/10/2024",
                "PONENTE",
                "Capacitación en Seguridad Vial",
                "https://sistema-antiguo.com/certificados/87654321.pdf",
                "Sistema Antiguo v1.0"
            ],
        ]
        
        for row_num, example in enumerate(examples, 2):
            for col_num, value in enumerate(example, 1):
                cell = ws.cell(row=row_num, column=col_num)
                cell.value = value
                cell.border = border
                cell.alignment = Alignment(horizontal='left', vertical='center')
                
                # Resaltar URLs
                if col_num == 6:  # URL del Certificado
                    cell.font = Font(color="0000FF", underline="single")
        
        # Ajustar anchos de columna
        ws.column_dimensions['A'].width = 12
        ws.column_dimensions['B'].width = 30
        ws.column_dimensions['C'].width = 18
        ws.column_dimensions['D'].width = 18
        ws.column_dimensions['E'].width = 40
        ws.column_dimensions['F'].width = 50
        ws.column_dimensions['G'].width = 25
        
        # Crear respuesta HTTP
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = 'attachment; filename=plantilla_certificados_externos.xlsx'
        
        wb.save(response)
        return response



# ============================================================================
# VISTAS PARA PROCESAMIENTO DE CERTIFICADOS CON QR
# ============================================================================

from django.views import View
from django.shortcuts import render, redirect
from django.http import HttpResponse, JsonResponse
from django.contrib.auth.mixins import LoginRequiredMixin, UserPassesTestMixin
from certificates.models import Certificate, Event, QRProcessingConfig
from certificates.services.pdf_processing import PDFProcessingService


class StaffRequiredMixin(UserPassesTestMixin):
    """Mixin para requerir que el usuario sea staff"""
    
    def test_func(self):
        return self.request.user.is_staff


@method_decorator(staff_member_required, name="dispatch")
class PDFImportView(StaffRequiredMixin, View):
    """Vista para importar certificados PDF originales"""
    
    template_name = "admin/certificates/pdf_import.html"
    
    def get(self, request):
        """Muestra el formulario de importación"""
        events = Event.objects.all().order_by('-event_date')
        context = {
            'events': events,
            'title': 'Importar Certificados PDF',
        }
        return render(request, self.template_name, context)
    
    def post(self, request):
        """Procesa la importación de PDFs"""
        pdf_files = request.FILES.getlist('pdf_files')
        event_id = request.POST.get('event_id')
        auto_extract_names = request.POST.get('auto_extract_names') == 'on'
        
        if not pdf_files:
            messages.error(request, '❌ No se seleccionaron archivos PDF')
            return redirect('certificates:pdf_import')
        
        if not event_id:
            messages.error(request, '❌ Debe seleccionar un evento')
            return redirect('certificates:pdf_import')
        
        try:
            event = Event.objects.get(id=event_id)
        except Event.DoesNotExist:
            messages.error(request, '❌ Evento no encontrado')
            return redirect('certificates:pdf_import')
        
        # Procesar PDFs
        service = PDFProcessingService()
        result = service.import_pdf_batch(
            pdf_files=pdf_files,
            event=event,
            auto_extract_names=auto_extract_names
        )
        
        # Mostrar resultados
        if result['success_count'] > 0:
            messages.success(
                request,
                f"✓ Se importaron {result['success_count']} certificados exitosamente"
            )
        
        if result['error_count'] > 0:
            messages.warning(
                request,
                f"⚠ Se encontraron {result['error_count']} errores"
            )
            for error in result['errors'][:5]:
                messages.error(request, error)
        
        if result['warnings']:
            for warning in result['warnings'][:5]:
                messages.warning(request, warning)
        
        return redirect('admin:certificates_certificate_changelist')


@method_decorator(staff_member_required, name="dispatch")
class QRProcessingView(StaffRequiredMixin, View):
    """Vista para procesar códigos QR en lote"""
    
    def post(self, request):
        """Procesa QR para certificados seleccionados"""
        certificate_ids = request.POST.getlist('certificate_ids')
        
        if not certificate_ids:
            messages.error(request, '❌ No se seleccionaron certificados')
            return redirect('admin:certificates_certificate_changelist')
        
        # Obtener configuración
        config = QRProcessingConfig.get_active_config()
        
        # Procesar cada certificado
        service = PDFProcessingService()
        success_count = 0
        error_count = 0
        errors = []
        
        for cert_id in certificate_ids:
            try:
                certificate = Certificate.objects.get(id=cert_id)
                result = service.process_qr_for_certificate(certificate, config)
                
                if result['success']:
                    success_count += 1
                else:
                    error_count += 1
                    errors.append(f"{certificate.participant.full_name}: {result['error']}")
            except Certificate.DoesNotExist:
                error_count += 1
                errors.append(f"Certificado ID {cert_id} no encontrado")
            except Exception as e:
                error_count += 1
                errors.append(f"Error en certificado ID {cert_id}: {str(e)}")
        
        # Mostrar resultados
        if success_count > 0:
            messages.success(
                request,
                f"✓ Se procesaron {success_count} certificados exitosamente"
            )
        
        if error_count > 0:
            messages.warning(
                request,
                f"⚠ Se encontraron {error_count} errores"
            )
            for error in errors[:5]:
                messages.error(request, error)
        
        return redirect('admin:certificates_certificate_changelist')


@method_decorator(staff_member_required, name="dispatch")
class ExportForSigningView(StaffRequiredMixin, View):
    """Vista para exportar certificados para firma digital"""
    
    def post(self, request):
        """Crea y descarga ZIP con certificados"""
        certificate_ids = request.POST.getlist('certificate_ids')
        
        if not certificate_ids:
            messages.error(request, '❌ No se seleccionaron certificados')
            return redirect('admin:certificates_certificate_changelist')
        
        # Obtener certificados
        certificates = Certificate.objects.filter(
            id__in=certificate_ids,
            processing_status='QR_INSERTED'
        )
        
        if not certificates.exists():
            messages.error(
                request,
                '❌ No hay certificados válidos para exportar (deben estar en estado QR_INSERTED)'
            )
            return redirect('admin:certificates_certificate_changelist')
        
        # Crear ZIP
        service = PDFProcessingService()
        try:
            zip_bytes, zip_filename = service.create_export_zip(
                certificates=list(certificates),
                include_metadata=True
            )
            
            # Retornar ZIP como descarga
            response = HttpResponse(zip_bytes, content_type='application/zip')
            response['Content-Disposition'] = f'attachment; filename="{zip_filename}"'
            
            messages.success(
                request,
                f"✓ Se exportaron {certificates.count()} certificados"
            )
            
            return response
            
        except Exception as e:
            messages.error(request, f"❌ Error al crear ZIP: {str(e)}")
            return redirect('admin:certificates_certificate_changelist')


@method_decorator(staff_member_required, name="dispatch")
class FinalImportView(StaffRequiredMixin, View):
    """Vista para importar certificados firmados finales"""
    
    template_name = "admin/certificates/final_import.html"
    
    def get(self, request):
        """Muestra el formulario de importación"""
        context = {
            'title': 'Importar Certificados Firmados',
        }
        return render(request, self.template_name, context)
    
    def post(self, request):
        """Procesa certificados firmados finales"""
        pdf_files = request.FILES.getlist('pdf_files')
        
        if not pdf_files:
            messages.error(request, '❌ No se seleccionaron archivos PDF')
            return redirect('certificates:final_import')
        
        # Procesar PDFs finales
        service = PDFProcessingService()
        result = service.import_final_certificates(pdf_files)
        
        # Mostrar resultados
        if result['success_count'] > 0:
            messages.success(
                request,
                f"✓ Se importaron {result['success_count']} certificados firmados exitosamente"
            )
        
        if result['error_count'] > 0:
            messages.warning(
                request,
                f"⚠ Se encontraron {result['error_count']} errores"
            )
            for error in result['errors'][:5]:
                messages.error(request, error)
        
        return redirect('admin:certificates_certificate_changelist')


@method_decorator(staff_member_required, name="dispatch")
class ProcessingStatusView(StaffRequiredMixin, View):
    """Vista del panel de estado de procesamiento"""
    
    template_name = "admin/certificates/processing_status.html"
    
    def get(self, request):
        """Muestra el panel de estado"""
        status_filter = request.GET.get('status', 'all')
        
        # Obtener certificados
        certificates = Certificate.objects.select_related(
            'participant', 'participant__event'
        ).order_by('-processed_at', '-created_at')
        
        # Filtrar por estado si se especifica
        if status_filter != 'all':
            certificates = certificates.filter(processing_status=status_filter)
        
        # Estadísticas
        stats = {
            'total': Certificate.objects.count(),
            'imported': Certificate.objects.filter(processing_status='IMPORTED').count(),
            'qr_inserted': Certificate.objects.filter(processing_status='QR_INSERTED').count(),
            'exported': Certificate.objects.filter(processing_status='EXPORTED_FOR_SIGNING').count(),
            'signed_final': Certificate.objects.filter(processing_status='SIGNED_FINAL').count(),
            'errors': Certificate.objects.filter(processing_status='ERROR').count(),
        }
        
        context = {
            'title': 'Estado de Procesamiento',
            'certificates': certificates[:100],  # Limitar a 100 para performance
            'stats': stats,
            'current_filter': status_filter,
        }
        
        return render(request, self.template_name, context)
