"""Tests para vistas de administración"""
from django.test import TestCase, Client, override_settings
from django.contrib.auth.models import User
from openpyxl import Workbook
from io import BytesIO
from datetime import date
from certificates.models import Participant


@override_settings(LOGIN_URL="/admin/login/")
class ExcelImportViewTest(TestCase):
    """Tests para la vista de importación de Excel"""

    def setUp(self):
        # Crear usuario staff
        self.staff_user = User.objects.create_user(
            username="staff", password="staff123", is_staff=True
        )

        # Crear usuario normal
        self.normal_user = User.objects.create_user(
            username="user", password="user123"
        )

        self.client = Client()
        self.url = "/admin/import-excel/"

    def _create_excel_file(self, data_rows):
        """Helper para crear archivo Excel de prueba"""
        workbook = Workbook()
        sheet = workbook.active

        # Encabezados
        headers = [
            "DNI",
            "Nombres y Apellidos",
            "Fecha del Evento",
            "Tipo de Asistente",
            "Nombre del Evento",
        ]
        for col_idx, header in enumerate(headers, start=1):
            sheet.cell(row=1, column=col_idx, value=header)

        # Datos
        for row_idx, row_data in enumerate(data_rows, start=2):
            for col_idx, value in enumerate(row_data, start=1):
                sheet.cell(row=row_idx, column=col_idx, value=value)

        # Guardar en BytesIO
        excel_buffer = BytesIO()
        workbook.save(excel_buffer)
        excel_buffer.seek(0)

        return excel_buffer

    def test_view_requires_staff_permission(self):
        """Solo usuarios staff pueden acceder"""
        # Usuario no autenticado
        response = self.client.get(self.url)
        self.assertEqual(response.status_code, 302)  # Redirect to login

        # Usuario normal (no staff)
        self.client.login(username="user", password="user123")
        response = self.client.get(self.url)
        self.assertEqual(response.status_code, 302)  # Redirect

        # Usuario staff
        self.client.logout()
        self.client.login(username="staff", password="staff123")
        response = self.client.get(self.url)
        self.assertEqual(response.status_code, 200)

    def test_view_displays_form(self):
        """Debe mostrar el formulario de importación"""
        self.client.login(username="staff", password="staff123")

        response = self.client.get(self.url)

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, "Importar Participantes desde Excel")
        self.assertContains(response, "excel_file")
        self.assertContains(response, "Instrucciones")

    def test_import_valid_excel(self):
        """Debe importar correctamente un archivo Excel válido"""
        from django.core.files.uploadedfile import SimpleUploadedFile
        from django.test import override_settings
        
        self.client.login(username="staff", password="staff123")

        # Crear archivo Excel válido
        data = [
            ["12345678", "Juan Pérez", date(2024, 1, 1), "ASISTENTE", "Capacitación Django"],
            ["87654321", "María López", date(2024, 1, 1), "PONENTE", "Capacitación Django"],
        ]
        excel_buffer = self._create_excel_file(data)
        excel_file = SimpleUploadedFile(
            "test.xlsx",
            excel_buffer.read(),
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )

        # Enviar formulario (el cliente de test maneja CSRF automáticamente)
        response = self.client.post(
            self.url, {"excel_file": excel_file}
        )

        # Verificar redirección
        self.assertEqual(response.status_code, 302)

        # Verificar que se crearon los participantes
        self.assertEqual(Participant.objects.count(), 2)
        self.assertTrue(Participant.objects.filter(dni="12345678").exists())
        self.assertTrue(Participant.objects.filter(dni="87654321").exists())

    def test_import_invalid_file_format(self):
        """Debe rechazar archivos con formato inválido"""
        self.client.login(username="staff", password="staff123")

        # Crear archivo de texto
        invalid_file = BytesIO(b"This is not an Excel file")
        invalid_file.name = "test.txt"

        # Enviar formulario
        response = self.client.post(
            self.url, {"excel_file": invalid_file}
        )

        # Verificar que muestra error de validación
        self.assertEqual(response.status_code, 200)
        self.assertContains(response, "Formato de archivo no válido")
