"""Tests de integración completos para el sistema de certificados"""
from django.test import TestCase, Client
from django.urls import reverse
from django.contrib.auth.models import User
from django.core.files.uploadedfile import SimpleUploadedFile
from openpyxl import Workbook
from io import BytesIO
from datetime import date, datetime
from unittest.mock import patch, MagicMock
import uuid as uuid_lib

from certificates.models import (
    Event, Participant, Certificate, CertificateTemplate, AuditLog
)
from certificates.services.excel_processor import ExcelProcessorService
from certificates.services.certificate_generator import CertificateGeneratorService
from certificates.services.digital_signature import DigitalSignatureService


class FullWorkflowIntegrationTest(TestCase):
    """Test del flujo completo: importar → generar → firmar → consultar → verificar"""

    def setUp(self):
        """Configuración inicial"""
        self.client = Client()
        
        # Crear usuario administrador
        self.admin_user = User.objects.create_superuser(
            username='admin',
            email='admin@test.com',
            password='admin123'
        )
        
        # Crear plantilla por defecto
        self.template = CertificateTemplate.objects.create(
            name="Plantilla por defecto",
            html_template="""
            <html>
            <body>
                <h1>Certificado de Participación</h1>
                <p>Nombre: {{ full_name }}</p>
                <p>DNI: {{ dni }}</p>
                <p>Evento: {{ event_name }}</p>
                <p>Fecha: {{ event_date }}</p>
                <p>Tipo: {{ attendee_type }}</p>
            </body>
            </html>
            """,
            is_default=True
        )

    def _create_test_excel(self, num_participants=3):
        """Helper para crear un archivo Excel de prueba"""
        workbook = Workbook()
        sheet = workbook.active
        
        # Encabezados
        headers = [
            'DNI', 'Nombres y Apellidos', 'Fecha del Evento',
            'Tipo de Asistente', 'Nombre del Evento'
        ]
        for col_idx, header in enumerate(headers, start=1):
            sheet.cell(row=1, column=col_idx, value=header)
        
        # Datos de participantes
        event_date = date(2024, 1, 15)
        event_name = 'Capacitación Django Avanzado'
        
        participants_data = [
            ['12345678', 'Juan Pérez García', event_date, 'ASISTENTE', event_name],
            ['87654321', 'María López Silva', event_date, 'PONENTE', event_name],
            ['11223344', 'Carlos Ruiz Mendoza', event_date, 'ORGANIZADOR', event_name],
        ]
        
        for row_idx, row_data in enumerate(participants_data[:num_participants], start=2):
            for col_idx, value in enumerate(row_data, start=1):
                sheet.cell(row=row_idx, column=col_idx, value=value)
        
        # Guardar en BytesIO
        excel_file = BytesIO()
        workbook.save(excel_file)
        excel_file.seek(0)
        excel_file.name = 'participantes.xlsx'
        
        return excel_file

    def test_full_workflow_import_generate_sign_query_verify(self):
        """Test del flujo completo del sistema"""
        
        # ===== PASO 1: IMPORTAR PARTICIPANTES DESDE EXCEL =====
        excel_file = self._create_test_excel(num_participants=3)
        excel_processor = ExcelProcessorService()
        
        import_result = excel_processor.process_excel(excel_file, user=self.admin_user)
        
        # Verificar importación exitosa
        self.assertEqual(import_result['success_count'], 3)
        self.assertEqual(import_result['error_count'], 0)
        self.assertEqual(Participant.objects.count(), 3)
        self.assertEqual(Event.objects.count(), 1)
        
        # Verificar que se creó log de auditoría
        import_log = AuditLog.objects.filter(action_type='IMPORT').first()
        self.assertIsNotNone(import_log)
        self.assertEqual(import_log.user, self.admin_user)
        
        # ===== PASO 2: GENERAR CERTIFICADOS =====
        event = Event.objects.first()
        certificate_generator = CertificateGeneratorService()
        
        generation_result = certificate_generator.generate_bulk_certificates(
            event, user=self.admin_user
        )
        
        # Verificar generación exitosa
        self.assertEqual(generation_result['success_count'], 3)
        self.assertEqual(generation_result['error_count'], 0)
        self.assertEqual(Certificate.objects.count(), 3)
        
        # Verificar que cada certificado tiene PDF y QR
        for certificate in Certificate.objects.all():
            self.assertIsNotNone(certificate.pdf_file)
            self.assertIsNotNone(certificate.qr_code)
            self.assertIsNotNone(certificate.uuid)
            self.assertIsNotNone(certificate.verification_url)
            self.assertFalse(certificate.is_signed)
        
        # Verificar logs de auditoría
        generate_logs = AuditLog.objects.filter(action_type='GENERATE')
        self.assertEqual(generate_logs.count(), 3)

        # ===== PASO 3: FIRMAR CERTIFICADOS =====
        with patch('requests.post') as mock_post:
            # Simular respuesta exitosa del servicio de firma
            mock_response = MagicMock()
            mock_response.status_code = 200
            mock_response.content = b'%PDF-1.4 Signed PDF content'
            mock_post.return_value = mock_response
            
            signature_service = DigitalSignatureService()
            certificates = Certificate.objects.all()
            
            sign_result = signature_service.sign_bulk_certificates(
                certificates, user=self.admin_user
            )
            
            # Verificar firma exitosa
            self.assertEqual(sign_result['success_count'], 3)
            self.assertEqual(sign_result['error_count'], 0)
            
            # Verificar que todos los certificados están firmados
            for certificate in Certificate.objects.all():
                certificate.refresh_from_db()
                self.assertTrue(certificate.is_signed)
                self.assertIsNotNone(certificate.signed_at)
            
            # Verificar logs de auditoría
            sign_logs = AuditLog.objects.filter(action_type='SIGN')
            self.assertEqual(sign_logs.count(), 3)
        
        # ===== PASO 4: CONSULTAR CERTIFICADOS POR DNI =====
        response = self.client.post(
            reverse('certificates:query'),
            {'dni': '12345678'}
        )
        
        # Verificar consulta exitosa
        self.assertEqual(response.status_code, 200)
        self.assertIn('certificates', response.context)
        certificates_found = response.context['certificates']
        self.assertEqual(len(certificates_found), 1)
        self.assertEqual(certificates_found[0].participant.dni, '12345678')
        self.assertEqual(certificates_found[0].participant.full_name, 'Juan Pérez García')
        
        # Verificar log de auditoría de consulta
        query_log = AuditLog.objects.filter(action_type='QUERY').first()
        self.assertIsNotNone(query_log)
        self.assertEqual(query_log.metadata['dni'], '12345678')
        self.assertEqual(query_log.metadata['results_count'], 1)
        
        # ===== PASO 5: VERIFICAR CERTIFICADO POR QR =====
        certificate = certificates_found[0]
        response = self.client.get(
            reverse('certificates:verify', kwargs={'uuid': certificate.uuid})
        )
        
        # Verificar verificación exitosa
        self.assertEqual(response.status_code, 200)
        self.assertTemplateUsed(response, 'certificates/verify.html')
        self.assertEqual(response.context['certificate'], certificate)
        
        # Verificar que muestra estado de firma
        self.assertContains(response, 'FIRMADO DIGITALMENTE')
        self.assertContains(response, certificate.participant.full_name)
        self.assertContains(response, certificate.participant.dni)
        
        # Verificar log de auditoría de verificación
        verify_log = AuditLog.objects.filter(action_type='VERIFY').first()
        self.assertIsNotNone(verify_log)
        self.assertEqual(verify_log.metadata['certificate_uuid'], str(certificate.uuid))
        self.assertTrue(verify_log.metadata['is_signed'])
        
        # ===== PASO 6: DESCARGAR CERTIFICADO =====
        response = self.client.get(
            reverse('certificates:download', kwargs={'uuid': certificate.uuid})
        )
        
        # Verificar descarga exitosa
        self.assertEqual(response.status_code, 200)
        self.assertEqual(response['Content-Type'], 'application/pdf')
        self.assertIn('attachment', response['Content-Disposition'])
        
        # ===== VERIFICACIÓN FINAL: AUDITORÍA COMPLETA =====
        # Debe haber logs de todas las operaciones
        self.assertEqual(AuditLog.objects.filter(action_type='IMPORT').count(), 1)
        self.assertEqual(AuditLog.objects.filter(action_type='GENERATE').count(), 3)
        self.assertEqual(AuditLog.objects.filter(action_type='SIGN').count(), 3)
        self.assertEqual(AuditLog.objects.filter(action_type='QUERY').count(), 1)
        self.assertEqual(AuditLog.objects.filter(action_type='VERIFY').count(), 1)


class ExcelImportIntegrationTest(TestCase):
    """Tests de integración para importación de Excel con archivos reales"""

    def setUp(self):
        """Configuración inicial"""
        self.service = ExcelProcessorService()
        self.admin_user = User.objects.create_user(
            username='admin',
            password='admin123'
        )

    def _create_real_excel_file(self):
        """Crea un archivo Excel realista con múltiples participantes"""
        workbook = Workbook()
        sheet = workbook.active
        
        # Encabezados
        headers = [
            'DNI', 'Nombres y Apellidos', 'Fecha del Evento',
            'Tipo de Asistente', 'Nombre del Evento'
        ]
        for col_idx, header in enumerate(headers, start=1):
            sheet.cell(row=1, column=col_idx, value=header)
        
        # Datos realistas de múltiples eventos
        participants = [
            # Evento 1: Capacitación Django
            ['12345678', 'Juan Carlos Pérez García', date(2024, 1, 15), 'ASISTENTE', 'Capacitación Django Avanzado'],
            ['87654321', 'María Elena López Silva', date(2024, 1, 15), 'PONENTE', 'Capacitación Django Avanzado'],
            ['11223344', 'Carlos Alberto Ruiz Mendoza', date(2024, 1, 15), 'ORGANIZADOR', 'Capacitación Django Avanzado'],
            ['55667788', 'Ana Patricia Torres Vega', date(2024, 1, 15), 'ASISTENTE', 'Capacitación Django Avanzado'],
            ['99887766', 'Luis Fernando Gómez Quispe', date(2024, 1, 15), 'ASISTENTE', 'Capacitación Django Avanzado'],
            
            # Evento 2: Capacitación Python
            ['12345678', 'Juan Carlos Pérez García', date(2024, 2, 20), 'PONENTE', 'Capacitación Python para Análisis de Datos'],
            ['22334455', 'Rosa María Flores Mamani', date(2024, 2, 20), 'ASISTENTE', 'Capacitación Python para Análisis de Datos'],
            ['33445566', 'Pedro José Condori Apaza', date(2024, 2, 20), 'ASISTENTE', 'Capacitación Python para Análisis de Datos'],
            ['44556677', 'Carmen Lucía Huanca Pari', date(2024, 2, 20), 'ORGANIZADOR', 'Capacitación Python para Análisis de Datos'],
            
            # Evento 3: Taller de Seguridad Vial
            ['87654321', 'María Elena López Silva', date(2024, 3, 10), 'ORGANIZADOR', 'Taller de Seguridad Vial'],
            ['66778899', 'Jorge Luis Mamani Ccama', date(2024, 3, 10), 'ASISTENTE', 'Taller de Seguridad Vial'],
            ['77889900', 'Silvia Beatriz Quispe Chura', date(2024, 3, 10), 'ASISTENTE', 'Taller de Seguridad Vial'],
        ]
        
        for row_idx, row_data in enumerate(participants, start=2):
            for col_idx, value in enumerate(row_data, start=1):
                sheet.cell(row=row_idx, column=col_idx, value=value)
        
        # Guardar en BytesIO
        excel_file = BytesIO()
        workbook.save(excel_file)
        excel_file.seek(0)
        excel_file.name = 'participantes_completo.xlsx'
        
        return excel_file

    def test_import_real_excel_with_multiple_events(self):
        """Test de importación con archivo Excel realista con múltiples eventos"""
        excel_file = self._create_real_excel_file()
        
        result = self.service.process_excel(excel_file, user=self.admin_user)
        
        # Verificar importación exitosa
        self.assertEqual(result['success_count'], 12)
        self.assertEqual(result['error_count'], 0)
        
        # Verificar que se crearon 3 eventos
        self.assertEqual(Event.objects.count(), 3)
        
        # Verificar eventos específicos
        django_event = Event.objects.get(name='Capacitación Django Avanzado')
        self.assertEqual(django_event.event_date, date(2024, 1, 15))
        self.assertEqual(django_event.participants.count(), 5)
        
        python_event = Event.objects.get(name='Capacitación Python para Análisis de Datos')
        self.assertEqual(python_event.event_date, date(2024, 2, 20))
        self.assertEqual(python_event.participants.count(), 4)
        
        seguridad_event = Event.objects.get(name='Taller de Seguridad Vial')
        self.assertEqual(seguridad_event.event_date, date(2024, 3, 10))
        self.assertEqual(seguridad_event.participants.count(), 3)
        
        # Verificar participantes duplicados en diferentes eventos
        juan_participants = Participant.objects.filter(dni='12345678')
        self.assertEqual(juan_participants.count(), 2)  # En 2 eventos diferentes
        
        maria_participants = Participant.objects.filter(dni='87654321')
        self.assertEqual(maria_participants.count(), 2)  # En 2 eventos diferentes

    def test_import_excel_with_mixed_valid_invalid_data(self):
        """Test de importación con datos válidos e inválidos mezclados"""
        workbook = Workbook()
        sheet = workbook.active
        
        # Encabezados
        headers = [
            'DNI', 'Nombres y Apellidos', 'Fecha del Evento',
            'Tipo de Asistente', 'Nombre del Evento'
        ]
        for col_idx, header in enumerate(headers, start=1):
            sheet.cell(row=1, column=col_idx, value=header)
        
        # Datos mezclados (válidos e inválidos)
        participants = [
            ['12345678', 'Juan Pérez', date(2024, 1, 15), 'ASISTENTE', 'Capacitación Test'],  # Válido
            ['123', 'María López', date(2024, 1, 15), 'PONENTE', 'Capacitación Test'],  # DNI inválido
            ['87654321', 'Carlos Ruiz', date(2024, 1, 15), 'INVALIDO', 'Capacitación Test'],  # Tipo inválido
            ['11223344', '', date(2024, 1, 15), 'ASISTENTE', 'Capacitación Test'],  # Nombre vacío
            ['55667788', 'Ana Torres', date(2024, 1, 15), 'ORGANIZADOR', 'Capacitación Test'],  # Válido
            ['99887766', 'Luis Gómez', date(2024, 1, 15), 'ASISTENTE', ''],  # Evento vacío
            ['22334455', 'Rosa Flores', date(2024, 1, 15), 'ASISTENTE', 'Capacitación Test'],  # Válido
        ]
        
        for row_idx, row_data in enumerate(participants, start=2):
            for col_idx, value in enumerate(row_data, start=1):
                sheet.cell(row=row_idx, column=col_idx, value=value)
        
        excel_file = BytesIO()
        workbook.save(excel_file)
        excel_file.seek(0)
        
        result = self.service.process_excel(excel_file, user=self.admin_user)
        
        # Verificar que se procesaron correctamente los válidos
        self.assertEqual(result['success_count'], 3)
        self.assertEqual(result['error_count'], 4)
        self.assertEqual(len(result['errors']), 4)
        
        # Verificar que solo se crearon los participantes válidos
        self.assertEqual(Participant.objects.count(), 3)
        
        # Verificar que los errores contienen información útil
        for error in result['errors']:
            self.assertIn('fila', error.lower())

    def test_import_excel_updates_existing_participants(self):
        """Test que la importación actualiza participantes existentes"""
        # Primera importación
        workbook = Workbook()
        sheet = workbook.active
        
        headers = ['DNI', 'Nombres y Apellidos', 'Fecha del Evento', 'Tipo de Asistente', 'Nombre del Evento']
        for col_idx, header in enumerate(headers, start=1):
            sheet.cell(row=1, column=col_idx, value=header)
        
        data = [['12345678', 'Juan Pérez', date(2024, 1, 15), 'ASISTENTE', 'Capacitación Test']]
        for row_idx, row_data in enumerate(data, start=2):
            for col_idx, value in enumerate(row_data, start=1):
                sheet.cell(row=row_idx, column=col_idx, value=value)
        
        excel_file = BytesIO()
        workbook.save(excel_file)
        excel_file.seek(0)
        
        result1 = self.service.process_excel(excel_file, user=self.admin_user)
        self.assertEqual(result1['success_count'], 1)
        
        participant_before = Participant.objects.get(dni='12345678')
        self.assertEqual(participant_before.full_name, 'Juan Pérez')
        self.assertEqual(participant_before.attendee_type, 'ASISTENTE')
        
        # Segunda importación con datos actualizados
        workbook2 = Workbook()
        sheet2 = workbook2.active
        
        for col_idx, header in enumerate(headers, start=1):
            sheet2.cell(row=1, column=col_idx, value=header)
        
        data2 = [['12345678', 'Juan Carlos Pérez García', date(2024, 1, 15), 'PONENTE', 'Capacitación Test']]
        for row_idx, row_data in enumerate(data2, start=2):
            for col_idx, value in enumerate(row_data, start=1):
                sheet2.cell(row=row_idx, column=col_idx, value=value)
        
        excel_file2 = BytesIO()
        workbook2.save(excel_file2)
        excel_file2.seek(0)
        
        result2 = self.service.process_excel(excel_file2, user=self.admin_user)
        self.assertEqual(result2['success_count'], 1)
        
        # Verificar que se actualizó el participante existente
        self.assertEqual(Participant.objects.filter(dni='12345678').count(), 1)
        
        participant_after = Participant.objects.get(dni='12345678')
        self.assertEqual(participant_after.full_name, 'Juan Carlos Pérez García')
        self.assertEqual(participant_after.attendee_type, 'PONENTE')


class CertificateGenerationIntegrationTest(TestCase):
    """Tests de integración para generación de certificados con plantillas reales"""

    def setUp(self):
        """Configuración inicial"""
        self.admin_user = User.objects.create_user(
            username='admin',
            password='admin123'
        )
        
        # Crear plantilla realista
        self.template = CertificateTemplate.objects.create(
            name="Plantilla DRTC Puno",
            html_template="""
            <!DOCTYPE html>
            <html>
            <head>
                <meta charset="UTF-8">
                <style>
                    @page {
                        size: A4 landscape;
                        margin: 2cm;
                    }
                    body {
                        font-family: Arial, sans-serif;
                        text-align: center;
                    }
                    .header {
                        font-size: 24px;
                        font-weight: bold;
                        color: #003366;
                        margin-bottom: 20px;
                    }
                    .title {
                        font-size: 36px;
                        font-weight: bold;
                        color: #0066cc;
                        margin: 30px 0;
                    }
                    .content {
                        font-size: 18px;
                        line-height: 1.8;
                        margin: 20px 0;
                    }
                    .name {
                        font-size: 28px;
                        font-weight: bold;
                        color: #000;
                        margin: 20px 0;
                    }
                    .footer {
                        margin-top: 50px;
                        font-size: 14px;
                    }
                </style>
            </head>
            <body>
                <div class="header">
                    DIRECCIÓN REGIONAL DE TRANSPORTES Y COMUNICACIONES PUNO
                </div>
                <div class="title">
                    CERTIFICADO DE PARTICIPACIÓN
                </div>
                <div class="content">
                    Se otorga el presente certificado a:
                </div>
                <div class="name">
                    {{ full_name }}
                </div>
                <div class="content">
                    Con DNI N° {{ dni }}<br>
                    Por su participación como <strong>{{ attendee_type }}</strong><br>
                    en el evento:<br>
                    <strong>{{ event_name }}</strong><br>
                    Realizado el {{ event_date }}
                </div>
                <div class="footer">
                    Puno, Perú
                </div>
            </body>
            </html>
            """,
            css_styles="",
            is_default=True
        )
        
        # Crear evento
        self.event = Event.objects.create(
            name="Capacitación en Normativa de Transporte Terrestre",
            event_date=date(2024, 1, 15),
            description="Capacitación sobre normativa vigente en transporte terrestre",
            template=self.template
        )
        
        # Crear participantes
        self.participants = [
            Participant.objects.create(
                dni="12345678",
                full_name="Juan Carlos Pérez García",
                event=self.event,
                attendee_type="ASISTENTE"
            ),
            Participant.objects.create(
                dni="87654321",
                full_name="María Elena López Silva",
                event=self.event,
                attendee_type="PONENTE"
            ),
            Participant.objects.create(
                dni="11223344",
                full_name="Carlos Alberto Ruiz Mendoza",
                event=self.event,
                attendee_type="ORGANIZADOR"
            ),
        ]

    def test_generate_certificates_with_real_template(self):
        """Test de generación de certificados con plantilla realista"""
        generator = CertificateGeneratorService()
        
        result = generator.generate_bulk_certificates(self.event, user=self.admin_user)
        
        # Verificar generación exitosa
        self.assertEqual(result['success_count'], 3)
        self.assertEqual(result['error_count'], 0)
        self.assertEqual(len(result['certificates']), 3)
        
        # Verificar cada certificado generado
        for certificate in result['certificates']:
            self.assertIsNotNone(certificate.pdf_file)
            self.assertIsNotNone(certificate.qr_code)
            self.assertIsNotNone(certificate.uuid)
            self.assertIsNotNone(certificate.verification_url)
            self.assertFalse(certificate.is_signed)
            
            # Verificar que el PDF existe y tiene contenido
            self.assertTrue(certificate.pdf_file.name)
            certificate.pdf_file.open('rb')
            pdf_content = certificate.pdf_file.read()
            certificate.pdf_file.close()
            self.assertGreater(len(pdf_content), 0)
            self.assertTrue(pdf_content.startswith(b'%PDF'))
            
            # Verificar que el QR existe
            self.assertTrue(certificate.qr_code.name)
            certificate.qr_code.open('rb')
            qr_content = certificate.qr_code.read()
            certificate.qr_code.close()
            self.assertGreater(len(qr_content), 0)

    def test_generate_certificates_with_different_attendee_types(self):
        """Test que los certificados se generan correctamente para diferentes tipos de asistentes"""
        generator = CertificateGeneratorService()
        
        # Generar certificados
        result = generator.generate_bulk_certificates(self.event, user=self.admin_user)
        
        # Verificar que cada tipo de asistente tiene su certificado
        asistente_cert = Certificate.objects.get(participant__attendee_type='ASISTENTE')
        ponente_cert = Certificate.objects.get(participant__attendee_type='PONENTE')
        organizador_cert = Certificate.objects.get(participant__attendee_type='ORGANIZADOR')
        
        # Todos deben tener archivos generados
        self.assertIsNotNone(asistente_cert.pdf_file)
        self.assertIsNotNone(ponente_cert.pdf_file)
        self.assertIsNotNone(organizador_cert.pdf_file)
        
        # Cada uno debe tener UUID único
        self.assertNotEqual(asistente_cert.uuid, ponente_cert.uuid)
        self.assertNotEqual(ponente_cert.uuid, organizador_cert.uuid)
        self.assertNotEqual(asistente_cert.uuid, organizador_cert.uuid)

    def test_regenerate_certificate_keeps_same_uuid(self):
        """Test que regenerar un certificado mantiene el mismo UUID"""
        generator = CertificateGeneratorService()
        
        # Generar certificado inicial
        participant = self.participants[0]
        cert1 = generator.generate_certificate(participant, user=self.admin_user)
        original_uuid = cert1.uuid
        original_pdf = cert1.pdf_file.name
        
        # Regenerar certificado
        cert2 = generator.generate_certificate(participant, user=self.admin_user)
        
        # Debe ser el mismo certificado con el mismo UUID
        self.assertEqual(cert1.id, cert2.id)
        self.assertEqual(cert2.uuid, original_uuid)
        
        # Verificar que solo existe un certificado
        self.assertEqual(Certificate.objects.filter(participant=participant).count(), 1)

    def test_generate_certificates_creates_verification_urls(self):
        """Test que se generan URLs de verificación correctas"""
        generator = CertificateGeneratorService()
        
        result = generator.generate_bulk_certificates(self.event, user=self.admin_user)
        
        for certificate in result['certificates']:
            # Verificar que la URL contiene el UUID
            self.assertIn(str(certificate.uuid), certificate.verification_url)
            
            # Verificar que la URL es válida
            self.assertTrue(
                certificate.verification_url.startswith('http://') or
                certificate.verification_url.startswith('https://')
            )


class PublicQueryIntegrationTest(TestCase):
    """Tests de integración para consulta pública con múltiples certificados"""

    def setUp(self):
        """Configuración inicial"""
        self.client = Client()
        
        # Crear plantilla
        self.template = CertificateTemplate.objects.create(
            name="Plantilla Test",
            html_template="<html><body>Test</body></html>",
            is_default=True
        )
        
        # Crear múltiples eventos
        self.events = [
            Event.objects.create(
                name=f"Capacitación {i}",
                event_date=date(2024, i, 1),
                template=self.template
            )
            for i in range(1, 6)  # 5 eventos
        ]
        
        # Crear participante con certificados en múltiples eventos
        self.dni_multiple = "12345678"
        for event in self.events:
            participant = Participant.objects.create(
                dni=self.dni_multiple,
                full_name="Juan Pérez García",
                event=event,
                attendee_type="ASISTENTE"
            )
            Certificate.objects.create(
                participant=participant,
                pdf_file=SimpleUploadedFile(f"cert_{event.id}.pdf", b"PDF content"),
                qr_code=SimpleUploadedFile(f"qr_{event.id}.png", b"QR content"),
                verification_url=f"http://testserver/verificar/{uuid_lib.uuid4()}/"
            )
        
        # Crear otro participante con un solo certificado
        self.dni_single = "87654321"
        participant_single = Participant.objects.create(
            dni=self.dni_single,
            full_name="María López Silva",
            event=self.events[0],
            attendee_type="PONENTE"
        )
        Certificate.objects.create(
            participant=participant_single,
            pdf_file=SimpleUploadedFile("cert_single.pdf", b"PDF content"),
            qr_code=SimpleUploadedFile("qr_single.png", b"QR content"),
            verification_url=f"http://testserver/verificar/{uuid_lib.uuid4()}/"
        )

    def test_query_participant_with_multiple_certificates(self):
        """Test de consulta para participante con múltiples certificados"""
        response = self.client.post(
            reverse('certificates:query'),
            {'dni': self.dni_multiple}
        )
        
        self.assertEqual(response.status_code, 200)
        self.assertIn('certificates', response.context)
        
        certificates = list(response.context['certificates'])
        
        # Debe retornar todos los certificados
        self.assertEqual(len(certificates), 5)
        
        # Verificar que están ordenados por fecha descendente
        dates = [cert.participant.event.event_date for cert in certificates]
        self.assertEqual(dates, sorted(dates, reverse=True))
        
        # Verificar que todos pertenecen al mismo DNI
        for cert in certificates:
            self.assertEqual(cert.participant.dni, self.dni_multiple)

    def test_query_participant_with_single_certificate(self):
        """Test de consulta para participante con un solo certificado"""
        response = self.client.post(
            reverse('certificates:query'),
            {'dni': self.dni_single}
        )
        
        self.assertEqual(response.status_code, 200)
        certificates = list(response.context['certificates'])
        
        self.assertEqual(len(certificates), 1)
        self.assertEqual(certificates[0].participant.dni, self.dni_single)
        self.assertEqual(certificates[0].participant.full_name, "María López Silva")

    def test_query_nonexistent_dni(self):
        """Test de consulta para DNI sin certificados"""
        response = self.client.post(
            reverse('certificates:query'),
            {'dni': '99999999'}
        )
        
        self.assertEqual(response.status_code, 200)
        certificates = list(response.context['certificates'])
        
        self.assertEqual(len(certificates), 0)

    def test_query_performance_with_many_certificates(self):
        """Test de performance con muchos certificados"""
        # Crear participante con muchos certificados
        dni_many = "11111111"
        
        for i in range(20):
            event = Event.objects.create(
                name=f"Evento {i}",
                event_date=date(2024, 1, i + 1),
                template=self.template
            )
            participant = Participant.objects.create(
                dni=dni_many,
                full_name="Test User",
                event=event,
                attendee_type="ASISTENTE"
            )
            Certificate.objects.create(
                participant=participant,
                pdf_file=SimpleUploadedFile(f"cert_{i}.pdf", b"PDF"),
                qr_code=SimpleUploadedFile(f"qr_{i}.png", b"QR"),
                verification_url=f"http://testserver/verificar/{uuid_lib.uuid4()}/"
            )
        
        # Consultar con límite de queries
        with self.assertNumQueries(2):  # 1 para certificados + 1 para audit log
            response = self.client.post(
                reverse('certificates:query'),
                {'dni': dni_many}
            )
            
            certificates = list(response.context['certificates'])
            self.assertEqual(len(certificates), 20)
            
            # Acceder a datos relacionados no debe generar queries adicionales
            for cert in certificates:
                _ = cert.participant.full_name
                _ = cert.participant.event.name

    def test_query_displays_all_certificate_info(self):
        """Test que la consulta muestra toda la información necesaria"""
        response = self.client.post(
            reverse('certificates:query'),
            {'dni': self.dni_multiple}
        )
        
        self.assertEqual(response.status_code, 200)
        
        # Verificar que se muestra información de cada certificado
        for event in self.events:
            self.assertContains(response, event.name)
        
        # Verificar que hay enlaces de descarga
        certificates = response.context['certificates']
        for cert in certificates:
            download_url = reverse('certificates:download', kwargs={'uuid': cert.uuid})
            self.assertContains(response, str(cert.uuid))


class QRVerificationIntegrationTest(TestCase):
    """Tests de integración para verificación con QR válido e inválido"""

    def setUp(self):
        """Configuración inicial"""
        self.client = Client()
        
        # Crear plantilla
        self.template = CertificateTemplate.objects.create(
            name="Plantilla Test",
            html_template="<html><body>Test</body></html>",
            is_default=True
        )
        
        # Crear evento
        self.event = Event.objects.create(
            name="Capacitación Test",
            event_date=date(2024, 1, 15),
            template=self.template
        )
        
        # Crear participante y certificado válido sin firma
        self.participant_unsigned = Participant.objects.create(
            dni="12345678",
            full_name="Juan Pérez García",
            event=self.event,
            attendee_type="ASISTENTE"
        )
        
        self.cert_unsigned = Certificate.objects.create(
            participant=self.participant_unsigned,
            pdf_file=SimpleUploadedFile("cert.pdf", b"PDF content"),
            qr_code=SimpleUploadedFile("qr.png", b"QR content"),
            verification_url=f"http://testserver/verificar/{uuid_lib.uuid4()}/",
            is_signed=False
        )
        
        # Crear participante y certificado válido con firma
        self.participant_signed = Participant.objects.create(
            dni="87654321",
            full_name="María López Silva",
            event=self.event,
            attendee_type="PONENTE"
        )
        
        self.cert_signed = Certificate.objects.create(
            participant=self.participant_signed,
            pdf_file=SimpleUploadedFile("cert2.pdf", b"PDF content signed"),
            qr_code=SimpleUploadedFile("qr2.png", b"QR content"),
            verification_url=f"http://testserver/verificar/{uuid_lib.uuid4()}/",
            is_signed=True,
            signed_at=datetime.now()
        )

    def test_verify_valid_qr_unsigned_certificate(self):
        """Test de verificación con QR válido de certificado sin firma"""
        response = self.client.get(
            reverse('certificates:verify', kwargs={'uuid': self.cert_unsigned.uuid})
        )
        
        self.assertEqual(response.status_code, 200)
        self.assertTemplateUsed(response, 'certificates/verify.html')
        
        # Verificar que muestra los datos correctos
        self.assertContains(response, self.participant_unsigned.dni)
        self.assertContains(response, self.participant_unsigned.full_name)
        self.assertContains(response, self.event.name)
        self.assertContains(response, 'Asistente')  # Display value is capitalized
        
        # Verificar que muestra estado sin firma
        self.assertContains(response, 'SIN FIRMA DIGITAL')
        
        # Verificar que se creó log de auditoría
        log = AuditLog.objects.filter(action_type='VERIFY').latest('timestamp')
        self.assertEqual(log.metadata['certificate_uuid'], str(self.cert_unsigned.uuid))
        self.assertFalse(log.metadata['is_signed'])

    def test_verify_valid_qr_signed_certificate(self):
        """Test de verificación con QR válido de certificado firmado"""
        response = self.client.get(
            reverse('certificates:verify', kwargs={'uuid': self.cert_signed.uuid})
        )
        
        self.assertEqual(response.status_code, 200)
        
        # Verificar que muestra los datos correctos
        self.assertContains(response, self.participant_signed.dni)
        self.assertContains(response, self.participant_signed.full_name)
        self.assertContains(response, self.event.name)
        self.assertContains(response, 'Ponente')  # Display value is capitalized
        
        # Verificar que muestra estado firmado
        self.assertContains(response, 'FIRMADO DIGITALMENTE')
        
        # Verificar que se creó log de auditoría
        log = AuditLog.objects.filter(action_type='VERIFY').latest('timestamp')
        self.assertEqual(log.metadata['certificate_uuid'], str(self.cert_signed.uuid))
        self.assertTrue(log.metadata['is_signed'])

    def test_verify_invalid_qr_uuid(self):
        """Test de verificación con UUID inválido"""
        invalid_uuid = uuid_lib.uuid4()
        
        response = self.client.get(
            reverse('certificates:verify', kwargs={'uuid': invalid_uuid})
        )
        
        self.assertEqual(response.status_code, 404)
        
        # Verificar que se creó log de auditoría del intento fallido
        log = AuditLog.objects.filter(action_type='VERIFY').latest('timestamp')
        self.assertEqual(log.metadata['certificate_uuid'], str(invalid_uuid))
        self.assertEqual(log.metadata['status'], 'not_found')

    def test_verify_malformed_uuid(self):
        """Test de verificación con UUID mal formado"""
        # Intentar con un UUID mal formado
        response = self.client.get('/verificar/invalid-uuid-format/')
        
        # Debe retornar 404 (la URL no coincide con el patrón)
        self.assertEqual(response.status_code, 404)

    def test_verify_tracks_ip_address(self):
        """Test que la verificación registra la IP del usuario"""
        response = self.client.get(
            reverse('certificates:verify', kwargs={'uuid': self.cert_unsigned.uuid}),
            REMOTE_ADDR='192.168.1.100'
        )
        
        self.assertEqual(response.status_code, 200)
        
        # Verificar que se registró la IP
        log = AuditLog.objects.filter(action_type='VERIFY').latest('timestamp')
        self.assertIsNotNone(log.ip_address)

    def test_verify_multiple_times_creates_multiple_logs(self):
        """Test que múltiples verificaciones crean múltiples logs"""
        initial_count = AuditLog.objects.filter(action_type='VERIFY').count()
        
        # Verificar 3 veces
        for _ in range(3):
            self.client.get(
                reverse('certificates:verify', kwargs={'uuid': self.cert_unsigned.uuid})
            )
        
        final_count = AuditLog.objects.filter(action_type='VERIFY').count()
        self.assertEqual(final_count, initial_count + 3)

    def test_verify_shows_event_date(self):
        """Test que la verificación muestra la fecha del evento"""
        response = self.client.get(
            reverse('certificates:verify', kwargs={'uuid': self.cert_unsigned.uuid})
        )
        
        self.assertEqual(response.status_code, 200)
        
        # Verificar que se muestra la fecha del evento
        self.assertContains(response, '15')  # Día
        self.assertContains(response, '2024')  # Año

    def test_verify_different_attendee_types(self):
        """Test de verificación para diferentes tipos de asistentes"""
        # Crear certificados para cada tipo
        types = ['ASISTENTE', 'PONENTE', 'ORGANIZADOR']
        
        for attendee_type in types:
            participant = Participant.objects.create(
                dni=f"1111111{types.index(attendee_type)}",
                full_name=f"Test {attendee_type}",
                event=self.event,
                attendee_type=attendee_type
            )
            
            certificate = Certificate.objects.create(
                participant=participant,
                pdf_file=SimpleUploadedFile(f"cert_{attendee_type}.pdf", b"PDF"),
                qr_code=SimpleUploadedFile(f"qr_{attendee_type}.png", b"QR"),
                verification_url=f"http://testserver/verificar/{uuid_lib.uuid4()}/",
                is_signed=False
            )
            
            response = self.client.get(
                reverse('certificates:verify', kwargs={'uuid': certificate.uuid})
            )
            
            self.assertEqual(response.status_code, 200)
            self.assertContains(response, attendee_type)
