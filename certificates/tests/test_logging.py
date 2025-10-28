"""Tests para verificar que el sistema de logging funciona correctamente"""
from django.test import TestCase, override_settings
from django.contrib.auth.models import User
from certificates.models import Event, Participant, Certificate, CertificateTemplate
from certificates.services.excel_processor import ExcelProcessorService
from certificates.services.certificate_generator import CertificateGeneratorService
from certificates.services.digital_signature import DigitalSignatureService
from datetime import date
import openpyxl
from io import BytesIO
import logging
from unittest.mock import patch, MagicMock


class LoggingTestCase(TestCase):
    """Tests para verificar que se generan logs correctamente"""

    def setUp(self):
        """Configurar datos de prueba"""
        self.user = User.objects.create_user(
            username='testuser',
            password='testpass123'
        )
        
        # Crear plantilla por defecto
        self.template = CertificateTemplate.objects.create(
            name='Plantilla Test',
            html_template='<html><body>{{ full_name }}</body></html>',
            is_default=True
        )
        
        # Crear evento
        self.event = Event.objects.create(
            name='Evento Test',
            event_date=date(2024, 1, 15),
            template=self.template
        )
        
        # Crear participante
        self.participant = Participant.objects.create(
            dni='12345678',
            full_name='Juan Pérez',
            event=self.event,
            attendee_type='ASISTENTE'
        )

    def _create_test_excel(self):
        """Crea un archivo Excel de prueba"""
        workbook = openpyxl.Workbook()
        sheet = workbook.active
        
        # Encabezados
        headers = ['DNI', 'Nombres y Apellidos', 'Fecha del Evento', 
                   'Tipo de Asistente', 'Nombre del Evento']
        sheet.append(headers)
        
        # Datos
        sheet.append(['87654321', 'María García', '15/01/2024', 'PONENTE', 'Evento Test'])
        
        # Guardar en BytesIO
        excel_buffer = BytesIO()
        workbook.save(excel_buffer)
        excel_buffer.seek(0)
        
        return excel_buffer

    @patch('certificates.services.excel_processor.logger')
    def test_excel_processor_logs_info(self, mock_logger):
        """Verifica que ExcelProcessorService registra logs de info"""
        service = ExcelProcessorService()
        excel_file = self._create_test_excel()
        
        # Procesar Excel
        result = service.process_excel(excel_file, self.user)
        
        # Verificar que se llamó al logger
        self.assertTrue(mock_logger.info.called)
        
        # Verificar que se registró la importación completada
        info_calls = [str(call) for call in mock_logger.info.call_args_list]
        self.assertTrue(
            any('Importación completada' in str(call) for call in info_calls),
            "Debe registrar log de importación completada"
        )

    @patch('certificates.services.excel_processor.logger')
    def test_excel_processor_logs_errors(self, mock_logger):
        """Verifica que ExcelProcessorService registra logs de errores"""
        service = ExcelProcessorService()
        
        # Crear archivo Excel inválido (sin columnas requeridas)
        workbook = openpyxl.Workbook()
        sheet = workbook.active
        sheet.append(['Columna1', 'Columna2'])
        
        excel_buffer = BytesIO()
        workbook.save(excel_buffer)
        excel_buffer.seek(0)
        
        # Procesar Excel inválido
        result = service.process_excel(excel_buffer, self.user)
        
        # Verificar que se llamó al logger de error
        self.assertTrue(mock_logger.error.called)
        
        # Verificar mensaje de error
        error_calls = [str(call) for call in mock_logger.error.call_args_list]
        self.assertTrue(
            any('Archivo inválido' in str(call) for call in error_calls),
            "Debe registrar log de archivo inválido"
        )

    @patch('certificates.services.excel_processor.logger')
    def test_excel_processor_logs_warnings(self, mock_logger):
        """Verifica que ExcelProcessorService registra logs de warnings"""
        service = ExcelProcessorService()
        
        # Crear archivo Excel con fila inválida
        workbook = openpyxl.Workbook()
        sheet = workbook.active
        
        headers = ['DNI', 'Nombres y Apellidos', 'Fecha del Evento', 
                   'Tipo de Asistente', 'Nombre del Evento']
        sheet.append(headers)
        
        # Fila con DNI inválido
        sheet.append(['123', 'Nombre Test', '15/01/2024', 'ASISTENTE', 'Evento Test'])
        
        excel_buffer = BytesIO()
        workbook.save(excel_buffer)
        excel_buffer.seek(0)
        
        # Procesar Excel
        result = service.process_excel(excel_buffer, self.user)
        
        # Verificar que se llamó al logger de warning
        self.assertTrue(mock_logger.warning.called)

    @patch('certificates.services.certificate_generator.logger')
    def test_certificate_generator_logs_info(self, mock_logger):
        """Verifica que CertificateGeneratorService registra logs de info"""
        service = CertificateGeneratorService()
        
        # Generar certificado
        certificate = service.generate_certificate(self.participant, self.user)
        
        # Verificar que se llamó al logger
        self.assertTrue(mock_logger.info.called)
        
        # Verificar que se registró la generación del certificado
        info_calls = [str(call) for call in mock_logger.info.call_args_list]
        self.assertTrue(
            any('Certificado generado' in str(call) for call in info_calls),
            "Debe registrar log de certificado generado"
        )

    def test_certificate_generator_logs_errors(self):
        """Verifica que CertificateGeneratorService registra logs de errores"""
        # Importar logging para capturar logs reales
        import logging
        from io import StringIO
        
        # Crear un handler de prueba para capturar logs
        log_stream = StringIO()
        test_handler = logging.StreamHandler(log_stream)
        test_handler.setLevel(logging.ERROR)
        
        # Obtener el logger y agregar el handler
        logger = logging.getLogger('certificates')
        logger.addHandler(test_handler)
        
        try:
            service = CertificateGeneratorService()
            
            # Crear un participante con datos que causarán error en PDF
            # Mockear el método _create_pdf para que lance una excepción
            with patch.object(service, '_create_pdf', side_effect=Exception("Error de prueba en PDF")):
                # Intentar generar certificado
                result = service.generate_bulk_certificates(self.event, self.user)
                
                # Verificar que hubo errores
                self.assertGreater(result['error_count'], 0, "Debe haber errores registrados")
                
                # Verificar que se registraron logs de error
                log_output = log_stream.getvalue()
                self.assertIn('Error al generar certificado', log_output, 
                             "Debe registrar log de error al generar certificado")
        finally:
            # Limpiar el handler
            logger.removeHandler(test_handler)

    @patch('certificates.services.certificate_generator.logger')
    def test_certificate_generator_bulk_logs(self, mock_logger):
        """Verifica que generate_bulk_certificates registra logs"""
        service = CertificateGeneratorService()
        
        # Crear más participantes
        Participant.objects.create(
            dni='11111111',
            full_name='Test User 1',
            event=self.event,
            attendee_type='ASISTENTE'
        )
        Participant.objects.create(
            dni='22222222',
            full_name='Test User 2',
            event=self.event,
            attendee_type='PONENTE'
        )
        
        # Generar certificados masivamente
        result = service.generate_bulk_certificates(self.event, self.user)
        
        # Verificar que se llamó al logger
        self.assertTrue(mock_logger.info.called)
        
        # Verificar que se registró la generación masiva
        info_calls = [str(call) for call in mock_logger.info.call_args_list]
        self.assertTrue(
            any('Generación masiva completada' in str(call) for call in info_calls),
            "Debe registrar log de generación masiva completada"
        )

    @patch('certificates.services.digital_signature.logger')
    @patch('certificates.services.digital_signature.requests.post')
    def test_digital_signature_logs_info(self, mock_post, mock_logger):
        """Verifica que DigitalSignatureService registra logs de info"""
        # Crear certificado
        service = CertificateGeneratorService()
        certificate = service.generate_certificate(self.participant, self.user)
        
        # Mock de respuesta exitosa del servicio de firma
        mock_response = MagicMock()
        mock_response.status_code = 200
        mock_response.content = b'PDF firmado'
        mock_post.return_value = mock_response
        
        # Firmar certificado
        signature_service = DigitalSignatureService()
        signature_service.sign_certificate(certificate)
        
        # Verificar que se llamó al logger
        self.assertTrue(mock_logger.info.called)
        
        # Verificar que se registró la firma exitosa
        info_calls = [str(call) for call in mock_logger.info.call_args_list]
        self.assertTrue(
            any('firmado exitosamente' in str(call) for call in info_calls),
            "Debe registrar log de firma exitosa"
        )

    @patch('certificates.services.digital_signature.logger')
    @patch('certificates.services.digital_signature.requests.post')
    def test_digital_signature_logs_debug(self, mock_post, mock_logger):
        """Verifica que DigitalSignatureService registra logs de debug"""
        # Crear certificado
        service = CertificateGeneratorService()
        certificate = service.generate_certificate(self.participant, self.user)
        
        # Mock de respuesta exitosa
        mock_response = MagicMock()
        mock_response.status_code = 200
        mock_response.content = b'PDF firmado'
        mock_post.return_value = mock_response
        
        # Firmar certificado
        signature_service = DigitalSignatureService()
        signature_service.sign_certificate(certificate)
        
        # Verificar que se llamó al logger de debug
        self.assertTrue(mock_logger.debug.called)

    @patch('certificates.services.digital_signature.logger')
    @patch('certificates.services.digital_signature.requests.post')
    def test_digital_signature_logs_errors(self, mock_post, mock_logger):
        """Verifica que DigitalSignatureService registra logs de errores"""
        # Crear certificado
        service = CertificateGeneratorService()
        certificate = service.generate_certificate(self.participant, self.user)
        
        # Mock de respuesta con error (timeout)
        mock_post.side_effect = requests.exceptions.Timeout("Connection timeout")
        
        # Intentar firmar certificado (debe fallar)
        signature_service = DigitalSignatureService()
        with self.assertRaises(Exception):
            signature_service.sign_certificate(certificate)
        
        # Verificar que se llamó al logger de error
        self.assertTrue(mock_logger.error.called)
        
        # Verificar mensaje de error
        error_calls = [str(call) for call in mock_logger.error.call_args_list]
        self.assertTrue(
            any('Falló la firma' in str(call) for call in error_calls),
            "Debe registrar log de error en firma"
        )

    @patch('certificates.services.digital_signature.logger')
    @patch('certificates.services.digital_signature.requests.post')
    def test_digital_signature_logs_warnings(self, mock_post, mock_logger):
        """Verifica que DigitalSignatureService registra logs de warnings"""
        # Crear certificado
        service = CertificateGeneratorService()
        certificate = service.generate_certificate(self.participant, self.user)
        
        # Mock de respuesta con error en primer intento, éxito en segundo
        mock_response_error = MagicMock()
        mock_response_error.status_code = 500
        mock_response_error.raise_for_status.side_effect = requests.exceptions.HTTPError("Server error")
        
        mock_response_success = MagicMock()
        mock_response_success.status_code = 200
        mock_response_success.content = b'PDF firmado'
        
        mock_post.side_effect = [mock_response_error, mock_response_success]
        
        # Firmar certificado (debe tener éxito en segundo intento)
        signature_service = DigitalSignatureService()
        signature_service.sign_certificate(certificate)
        
        # Verificar que se llamó al logger de warning
        self.assertTrue(mock_logger.warning.called)
        
        # Verificar mensaje de warning
        warning_calls = [str(call) for call in mock_logger.warning.call_args_list]
        self.assertTrue(
            any('Intento' in str(call) and 'falló' in str(call) for call in warning_calls),
            "Debe registrar log de warning en reintento"
        )

    @patch('certificates.services.digital_signature.logger')
    @patch('certificates.services.digital_signature.requests.post')
    def test_digital_signature_bulk_logs(self, mock_post, mock_logger):
        """Verifica que sign_bulk_certificates registra logs"""
        # Crear certificados
        service = CertificateGeneratorService()
        cert1 = service.generate_certificate(self.participant, self.user)
        
        participant2 = Participant.objects.create(
            dni='11111111',
            full_name='Test User 2',
            event=self.event,
            attendee_type='PONENTE'
        )
        cert2 = service.generate_certificate(participant2, self.user)
        
        # Mock de respuesta exitosa
        mock_response = MagicMock()
        mock_response.status_code = 200
        mock_response.content = b'PDF firmado'
        mock_post.return_value = mock_response
        
        # Firmar certificados masivamente
        signature_service = DigitalSignatureService()
        result = signature_service.sign_bulk_certificates([cert1, cert2])
        
        # Verificar que se llamó al logger
        self.assertTrue(mock_logger.info.called)
        
        # Verificar que se registró la firma masiva
        info_calls = [str(call) for call in mock_logger.info.call_args_list]
        self.assertTrue(
            any('Firma masiva completada' in str(call) for call in info_calls),
            "Debe registrar log de firma masiva completada"
        )

    def test_logging_configuration_exists(self):
        """Verifica que la configuración de logging existe en settings"""
        from django.conf import settings
        
        # Verificar que existe configuración de logging
        self.assertTrue(hasattr(settings, 'LOGGING'))
        self.assertIsInstance(settings.LOGGING, dict)
        
        # Verificar que existen los loggers requeridos
        loggers = settings.LOGGING.get('loggers', {})
        self.assertIn('certificates', loggers)
        self.assertIn('certificates.signature', loggers)
        
        # Verificar nivel de log para certificates
        self.assertEqual(loggers['certificates']['level'], 'INFO')
        
        # Verificar nivel de log para certificates.signature
        self.assertEqual(loggers['certificates.signature']['level'], 'DEBUG')

    def test_logging_handlers_configured(self):
        """Verifica que los handlers de logging están configurados"""
        from django.conf import settings
        
        handlers = settings.LOGGING.get('handlers', {})
        
        # Verificar que existe handler de archivo
        self.assertIn('file', handlers)
        self.assertEqual(handlers['file']['class'], 'logging.handlers.RotatingFileHandler')
        
        # Verificar que existe handler de archivo para firma
        self.assertIn('signature_file', handlers)
        self.assertEqual(handlers['signature_file']['class'], 'logging.handlers.RotatingFileHandler')
        
        # Verificar configuración de RotatingFileHandler
        self.assertEqual(handlers['file']['maxBytes'], 10485760)  # 10MB
        self.assertEqual(handlers['file']['backupCount'], 5)


# Importar requests para los tests
import requests
