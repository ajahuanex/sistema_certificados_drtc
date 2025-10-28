"""Tests para ExcelProcessorService"""
from django.test import TestCase
from openpyxl import Workbook
from io import BytesIO
from datetime import datetime
from certificates.services.excel_processor import ExcelProcessorService


class ExcelProcessorServiceTest(TestCase):
    """Tests para el servicio de procesamiento de Excel"""
    
    def setUp(self):
        self.service = ExcelProcessorService()
    
    def _create_test_excel(self, headers):
        """Helper para crear un archivo Excel de prueba"""
        workbook = Workbook()
        sheet = workbook.active
        
        # Agregar encabezados
        for col_idx, header in enumerate(headers, start=1):
            sheet.cell(row=1, column=col_idx, value=header)
        
        # Guardar en BytesIO
        excel_file = BytesIO()
        workbook.save(excel_file)
        excel_file.seek(0)
        
        return excel_file
    
    def test_validate_file_with_correct_columns(self):
        """Debe validar correctamente un archivo con todas las columnas requeridas"""
        headers = [
            'DNI',
            'Nombres y Apellidos',
            'Fecha del Evento',
            'Tipo de Asistente',
            'Nombre del Evento'
        ]
        excel_file = self._create_test_excel(headers)
        
        is_valid, errors = self.service.validate_file(excel_file)
        
        self.assertTrue(is_valid)
        self.assertEqual(len(errors), 0)
    
    def test_validate_file_with_missing_columns(self):
        """Debe detectar columnas faltantes"""
        headers = ['DNI', 'Nombres y Apellidos']  # Faltan columnas
        excel_file = self._create_test_excel(headers)
        
        is_valid, errors = self.service.validate_file(excel_file)
        
        self.assertFalse(is_valid)
        self.assertGreater(len(errors), 0)
        self.assertIn('Columnas faltantes', errors[0])
    
    def test_validate_file_with_extra_columns(self):
        """Debe validar correctamente aunque haya columnas extra"""
        headers = [
            'DNI',
            'Nombres y Apellidos',
            'Fecha del Evento',
            'Tipo de Asistente',
            'Nombre del Evento',
            'Columna Extra'  # Columna adicional
        ]
        excel_file = self._create_test_excel(headers)
        
        is_valid, errors = self.service.validate_file(excel_file)
        
        self.assertTrue(is_valid)
        self.assertEqual(len(errors), 0)

    
    def test_parse_row(self):
        """Debe parsear correctamente una fila"""
        headers = ['DNI', 'Nombres y Apellidos', 'Fecha del Evento']
        
        # Crear una fila mock
        class MockCell:
            def __init__(self, value):
                self.value = value
        
        row = [
            MockCell('12345678'),
            MockCell('Juan Pérez'),
            MockCell('01/01/2024')
        ]
        
        row_data = self.service._parse_row(row, headers)
        
        self.assertEqual(row_data['DNI'], '12345678')
        self.assertEqual(row_data['Nombres y Apellidos'], 'Juan Pérez')
        self.assertEqual(row_data['Fecha del Evento'], '01/01/2024')
    
    def test_validate_row_with_valid_data(self):
        """Debe validar correctamente una fila con datos válidos"""
        row_data = {
            'DNI': '12345678',
            'Nombres y Apellidos': 'Juan Pérez',
            'Fecha del Evento': '01/01/2024',
            'Tipo de Asistente': 'ASISTENTE',
            'Nombre del Evento': 'Capacitación Django'
        }
        
        is_valid, error = self.service._validate_row(row_data)
        
        self.assertTrue(is_valid)
        self.assertEqual(error, "")
    
    def test_validate_row_with_invalid_dni(self):
        """Debe rechazar DNI con formato inválido"""
        row_data = {
            'DNI': '123',  # DNI muy corto
            'Nombres y Apellidos': 'Juan Pérez',
            'Fecha del Evento': '01/01/2024',
            'Tipo de Asistente': 'ASISTENTE',
            'Nombre del Evento': 'Capacitación Django'
        }
        
        is_valid, error = self.service._validate_row(row_data)
        
        self.assertFalse(is_valid)
        self.assertIn('DNI', error)
        self.assertIn('8 dígitos', error)
    
    def test_validate_row_with_invalid_attendee_type(self):
        """Debe rechazar tipo de asistente inválido"""
        row_data = {
            'DNI': '12345678',
            'Nombres y Apellidos': 'Juan Pérez',
            'Fecha del Evento': '01/01/2024',
            'Tipo de Asistente': 'INVALIDO',
            'Nombre del Evento': 'Capacitación Django'
        }
        
        is_valid, error = self.service._validate_row(row_data)
        
        self.assertFalse(is_valid)
        self.assertIn('Tipo de asistente', error)
    
    def test_validate_row_with_empty_field(self):
        """Debe rechazar filas con campos vacíos"""
        row_data = {
            'DNI': '12345678',
            'Nombres y Apellidos': '',  # Campo vacío
            'Fecha del Evento': '01/01/2024',
            'Tipo de Asistente': 'ASISTENTE',
            'Nombre del Evento': 'Capacitación Django'
        }
        
        is_valid, error = self.service._validate_row(row_data)
        
        self.assertFalse(is_valid)
        self.assertIn('vacío', error)
    
    def test_validate_row_with_datetime_object(self):
        """Debe aceptar objetos datetime para la fecha"""
        row_data = {
            'DNI': '12345678',
            'Nombres y Apellidos': 'Juan Pérez',
            'Fecha del Evento': datetime(2024, 1, 1),
            'Tipo de Asistente': 'PONENTE',
            'Nombre del Evento': 'Capacitación Django'
        }
        
        is_valid, error = self.service._validate_row(row_data)
        
        self.assertTrue(is_valid)
        self.assertEqual(error, "")

    
    def test_create_or_update_participant_creates_new(self):
        """Debe crear un nuevo participante y evento"""
        row_data = {
            'DNI': '12345678',
            'Nombres y Apellidos': 'Juan Pérez',
            'Fecha del Evento': '01/01/2024',
            'Tipo de Asistente': 'ASISTENTE',
            'Nombre del Evento': 'Capacitación Django'
        }
        
        participant = self.service._create_or_update_participant(row_data)
        
        self.assertIsNotNone(participant)
        self.assertEqual(participant.dni, '12345678')
        self.assertEqual(participant.full_name, 'Juan Pérez')
        self.assertEqual(participant.attendee_type, 'ASISTENTE')
        self.assertEqual(participant.event.name, 'Capacitación Django')
    
    def test_create_or_update_participant_updates_existing(self):
        """Debe actualizar un participante existente"""
        from certificates.models import Event, Participant
        from datetime import date
        
        # Crear evento y participante inicial
        event = Event.objects.create(
            name='Capacitación Django',
            event_date=date(2024, 1, 1)
        )
        
        participant = Participant.objects.create(
            dni='12345678',
            full_name='Juan Pérez',
            event=event,
            attendee_type='ASISTENTE'
        )
        
        # Actualizar con nuevos datos
        row_data = {
            'DNI': '12345678',
            'Nombres y Apellidos': 'Juan Pérez García',  # Nombre actualizado
            'Fecha del Evento': '01/01/2024',
            'Tipo de Asistente': 'PONENTE',  # Tipo actualizado
            'Nombre del Evento': 'Capacitación Django'
        }
        
        updated_participant = self.service._create_or_update_participant(row_data)
        
        # Verificar que es el mismo participante actualizado
        self.assertEqual(updated_participant.id, participant.id)
        self.assertEqual(updated_participant.full_name, 'Juan Pérez García')
        self.assertEqual(updated_participant.attendee_type, 'PONENTE')
    
    def test_create_or_update_participant_reuses_event(self):
        """Debe reutilizar un evento existente"""
        from certificates.models import Event
        from datetime import date
        
        # Crear evento inicial
        event = Event.objects.create(
            name='Capacitación Django',
            event_date=date(2024, 1, 1)
        )
        
        row_data = {
            'DNI': '87654321',
            'Nombres y Apellidos': 'María López',
            'Fecha del Evento': '01/01/2024',
            'Tipo de Asistente': 'ASISTENTE',
            'Nombre del Evento': 'Capacitación Django'
        }
        
        participant = self.service._create_or_update_participant(row_data)
        
        # Verificar que usa el mismo evento
        self.assertEqual(participant.event.id, event.id)
        self.assertEqual(Event.objects.count(), 1)

    
    def test_process_excel_with_valid_data(self):
        """Debe procesar correctamente un archivo Excel válido"""
        from datetime import date
        
        # Crear archivo Excel con datos válidos
        workbook = Workbook()
        sheet = workbook.active
        
        # Encabezados
        headers = ['DNI', 'Nombres y Apellidos', 'Fecha del Evento', 'Tipo de Asistente', 'Nombre del Evento']
        for col_idx, header in enumerate(headers, start=1):
            sheet.cell(row=1, column=col_idx, value=header)
        
        # Datos
        data = [
            ['12345678', 'Juan Pérez', date(2024, 1, 1), 'ASISTENTE', 'Capacitación Django'],
            ['87654321', 'María López', date(2024, 1, 1), 'PONENTE', 'Capacitación Django'],
            ['11223344', 'Carlos Ruiz', date(2024, 1, 1), 'ORGANIZADOR', 'Capacitación Django'],
        ]
        
        for row_idx, row_data in enumerate(data, start=2):
            for col_idx, value in enumerate(row_data, start=1):
                sheet.cell(row=row_idx, column=col_idx, value=value)
        
        # Guardar en BytesIO
        excel_file = BytesIO()
        workbook.save(excel_file)
        excel_file.seek(0)
        
        # Procesar
        result = self.service.process_excel(excel_file)
        
        self.assertEqual(result['success_count'], 3)
        self.assertEqual(result['error_count'], 0)
        self.assertEqual(len(result['errors']), 0)
    
    def test_process_excel_with_invalid_rows(self):
        """Debe manejar filas con errores"""
        from datetime import date
        
        workbook = Workbook()
        sheet = workbook.active
        
        # Encabezados
        headers = ['DNI', 'Nombres y Apellidos', 'Fecha del Evento', 'Tipo de Asistente', 'Nombre del Evento']
        for col_idx, header in enumerate(headers, start=1):
            sheet.cell(row=1, column=col_idx, value=header)
        
        # Datos (algunos inválidos)
        data = [
            ['12345678', 'Juan Pérez', date(2024, 1, 1), 'ASISTENTE', 'Capacitación Django'],  # Válido
            ['123', 'María López', date(2024, 1, 1), 'PONENTE', 'Capacitación Django'],  # DNI inválido
            ['11223344', 'Carlos Ruiz', date(2024, 1, 1), 'INVALIDO', 'Capacitación Django'],  # Tipo inválido
        ]
        
        for row_idx, row_data in enumerate(data, start=2):
            for col_idx, value in enumerate(row_data, start=1):
                sheet.cell(row=row_idx, column=col_idx, value=value)
        
        excel_file = BytesIO()
        workbook.save(excel_file)
        excel_file.seek(0)
        
        result = self.service.process_excel(excel_file)
        
        self.assertEqual(result['success_count'], 1)
        self.assertEqual(result['error_count'], 2)
        self.assertGreater(len(result['errors']), 0)
    
    def test_process_excel_creates_audit_log(self):
        """Debe crear un registro de auditoría"""
        from certificates.models import AuditLog
        from datetime import date
        
        workbook = Workbook()
        sheet = workbook.active
        
        headers = ['DNI', 'Nombres y Apellidos', 'Fecha del Evento', 'Tipo de Asistente', 'Nombre del Evento']
        for col_idx, header in enumerate(headers, start=1):
            sheet.cell(row=1, column=col_idx, value=header)
        
        data = [['12345678', 'Juan Pérez', date(2024, 1, 1), 'ASISTENTE', 'Capacitación Django']]
        for row_idx, row_data in enumerate(data, start=2):
            for col_idx, value in enumerate(row_data, start=1):
                sheet.cell(row=row_idx, column=col_idx, value=value)
        
        excel_file = BytesIO()
        workbook.save(excel_file)
        excel_file.seek(0)
        
        # Verificar que no hay logs antes
        initial_count = AuditLog.objects.count()
        
        # Procesar
        self.service.process_excel(excel_file)
        
        # Verificar que se creó un log
        self.assertEqual(AuditLog.objects.count(), initial_count + 1)
        
        log = AuditLog.objects.latest('timestamp')
        self.assertEqual(log.action_type, 'IMPORT')
        self.assertIn('success_count', log.metadata)
